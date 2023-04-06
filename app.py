from flask import Flask, render_template, request, jsonify
import openpyxl

from chat import get_response

app = Flask(__name__)

@app.get("/")
def index_get():
    return render_template('base.html')

@app.post("/predict")
def predict():
    text = request.get_json().get("message")
    response = get_response(text)
    
    # if response is a list of default options, return them as a string
    if isinstance(response, list):
        message = {"answer":"\n".join(response)}
    else:
        message = {"answer":response}
    
    # write the response to Conversations.xlsx file
    wb = openpyxl.load_workbook('Conversations.xlsx')
    sheet = wb.active
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1, value=text)
    sheet.cell(row=next_row, column=2, value=response)
    wb.save('Conversations.xlsx')
     
    return jsonify(message)

if __name__ == "__main__":
    app.run(debug=True)
